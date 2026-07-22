from pathlib import Path
import re

from openpyxl import load_workbook as _load_workbook
from robot.api.deco import keyword
from robot.libraries.BuiltIn import BuiltIn


ROBOT_LIBRARY_SCOPE = "GLOBAL"
ROBOT_AUTO_KEYWORDS = False

RUTA_EXCEL_DEFAULT = Path(__file__).resolve().parent.parent.parent / "variables" / "variables.xlsx"


@keyword("Cargar variables de excel")
def cargar_variables_de_excel(codigo_caso=None, fila=1, archivo=None):
    codigo = _normalizar_codigo(codigo_caso or _codigo_desde_tags())
    ruta_excel = Path(archivo).resolve() if archivo else RUTA_EXCEL_DEFAULT
    variables = _leer_variables(ruta_excel, codigo, int(fila))

    built_in = BuiltIn()
    built_in.set_test_variable("${codigo_caso}", codigo)
    built_in.set_test_variable("${fila_datos}", int(fila))

    for nombre, valor in variables.items():
        built_in.set_test_variable("${" + nombre + "}", valor)

    built_in.log(f"Variables cargadas desde {ruta_excel.name}, hoja {codigo}, fila de datos {fila}")
    return codigo


@keyword("Obtener filas con datos")
def obtener_filas_con_datos(codigo_caso, archivo=None):
    codigo = _normalizar_codigo(codigo_caso)
    ruta_excel = Path(archivo).resolve() if archivo else RUTA_EXCEL_DEFAULT
    return _leer_filas_con_datos(ruta_excel, codigo)


@keyword("Obtener codigo caso actual")
def obtener_codigo_caso_actual():
    return _normalizar_codigo(_codigo_desde_tags())


def _leer_variables(ruta_excel, hoja, fila_datos):
    libro = _load_workbook(ruta_excel, data_only=True, read_only=True)
    try:
        if hoja not in libro.sheetnames:
            raise ValueError(f"No existe la hoja {hoja} en {ruta_excel}")

        filas = list(libro[hoja].iter_rows(values_only=True))
    finally:
        libro.close()

    if not filas:
        raise ValueError(f"La hoja {hoja} esta vacia")

    encabezados = [_normalizar_variable(celda) for celda in filas[0] if celda]
    if not encabezados:
        raise ValueError(f"La hoja {hoja} no tiene encabezados")

    indice_fila = fila_datos
    if indice_fila <= 0 or indice_fila >= len(filas):
        raise ValueError(f"No existe la fila de datos {fila_datos} en la hoja {hoja}")

    valores = filas[indice_fila]
    return {
        encabezado: valores[indice] if indice < len(valores) and valores[indice] is not None else ""
        for indice, encabezado in enumerate(encabezados)
    }


def _leer_filas_con_datos(ruta_excel, hoja):
    libro = _load_workbook(ruta_excel, data_only=True, read_only=True)
    try:
        if hoja not in libro.sheetnames:
            raise ValueError(f"No existe la hoja {hoja} en {ruta_excel}")

        filas = list(libro[hoja].iter_rows(values_only=True))
    finally:
        libro.close()

    if not filas:
        raise ValueError(f"La hoja {hoja} esta vacia")

    encabezados = [_normalizar_variable(celda) for celda in filas[0] if celda]
    if not encabezados:
        raise ValueError(f"La hoja {hoja} no tiene encabezados")

    filas_con_datos = []
    for indice_fila, valores in enumerate(filas[1:], start=1):
        valores_variables = [
            valores[indice] if indice < len(valores) else None
            for indice in range(len(encabezados))
        ]
        if any(_tiene_dato(valor) for valor in valores_variables):
            filas_con_datos.append(indice_fila)

    return filas_con_datos


def _codigo_desde_tags():
    tags = BuiltIn().get_variable_value("@{TEST TAGS}", [])
    for tag in tags:
        if re.fullmatch(r"TC[-_]?\d{1,4}", str(tag).strip(), flags=re.IGNORECASE):
            return tag
    raise ValueError("El caso debe tener un tag tipo TC-001 para cargar datos desde Excel")


def _normalizar_codigo(codigo):
    valor = str(codigo).strip().upper().replace("_", "-")
    if valor.isdigit():
        return f"TC-{int(valor):03}"

    match = re.fullmatch(r"TC-?(\d{1,4})", valor)
    if match:
        return f"TC-{int(match.group(1)):03}"

    return valor


def _normalizar_variable(nombre):
    variable = str(nombre).strip().lower()
    variable = re.sub(r"\s+", "_", variable)
    variable = re.sub(r"[^a-z0-9_]", "_", variable)
    variable = re.sub(r"_+", "_", variable).strip("_")
    if not variable:
        raise ValueError("Hay un encabezado vacio o invalido en el Excel")
    return variable


def _tiene_dato(valor):
    return valor is not None and str(valor).strip() != ""

from pathlib import Path
import re

from openpyxl import load_workbook


RUTA_EXCEL = Path(__file__).with_name("variables.xlsx")


def get_variables(testcase="TC-001", fila=1, archivo=None):
    hoja = _normalizar_hoja(testcase)
    fila = int(fila)
    ruta_excel = Path(archivo) if archivo else RUTA_EXCEL

    libro = load_workbook(ruta_excel, data_only=True, read_only=True)
    if hoja not in libro.sheetnames:
        raise ValueError(f"No existe la hoja {hoja} en {ruta_excel}")

    filas = list(libro[hoja].iter_rows(values_only=True))
    if not filas:
        raise ValueError(f"La hoja {hoja} esta vacia")

    columnas = [_normalizar_variable(columna) for columna in filas[0] if columna]
    if not columnas:
        raise ValueError(f"La hoja {hoja} no tiene encabezados")

    indice_fila = fila
    if indice_fila >= len(filas):
        raise ValueError(f"No existe la fila de datos {fila} en la hoja {hoja}")

    valores = filas[indice_fila]
    return {
        columna: valores[indice] if indice < len(valores) and valores[indice] is not None else ""
        for indice, columna in enumerate(columnas)
    }


def _normalizar_hoja(testcase):
    valor = str(testcase).strip().upper().replace("_", "-")
    if valor.isdigit():
        return f"TC-{int(valor):03}"

    match = re.fullmatch(r"TC-?(\d{1,4})", valor)
    if match:
        return f"TC-{int(match.group(1)):03}"

    return valor


def _normalizar_variable(nombre):
    variable = str(nombre).strip().lower()
    variable = re.sub(r"\s+", "_", variable)
    variable = re.sub(r"[^a-z0-9_]", "_", variable)
    variable = re.sub(r"_+", "_", variable).strip("_")
    if not variable:
        raise ValueError("Hay un encabezado vacio o invalido en el Excel")
    return variable
